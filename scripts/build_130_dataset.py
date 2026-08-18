import openpyxl
import re
import json
import os
import datetime

excel_path = 'C:/Users/lipca/Downloads/UZEE_TECH_SUPER_D_MASTER_130_SCREEN_GUARD_FINAL.xlsx'
wb = openpyxl.load_workbook(excel_path)
sheet = wb['SCREEN_GUARD_130']

# Abbreviation Map
BRAND_ABBREVIATIONS = {
    'IP': 'iPhone',
    'IPHONE': 'iPhone',
    'SAM': 'Samsung',
    'SAMSUNG': 'Samsung',
    'RM': 'Redmi',
    'REDMI': 'Redmi',
    'OP': 'OPPO',
    'OPPO': 'OPPO',
    'VO': 'Vivo',
    'VIVO': 'Vivo',
    '1+': 'OnePlus',
    'ONEPLUS': 'OnePlus',
    'POC': 'POCO',
    'POCO': 'POCO',
    'XM': 'Xiaomi',
    'XIAOMI': 'Xiaomi',
    'REAL': 'Realme',
    'REALME': 'Realme',
    'TECNO': 'Tecno',
    'HUAWEI': 'Huawei',
    'HONOR': 'Honor',
    'INFINIX': 'Infinix',
    'MOTO': 'Motorola',
    'MOTOROLA': 'Motorola',
    'NOKIA': 'Nokia',
    'ZTE': 'ZTE',
    'PIXEL': 'Google Pixel'
}

# Display sizes per Section 6 & 7 of prompt
DISPLAY_SIZES = {
    1: '4.7"',
    2: '5.5"',
    3: '5.8"',
    4: '6.1"',
    5: '6.5"',
    6: '6.1"',
    7: '5.4"',
    8: '6.7"',
    9: '6.1"',
    10: '6.7"',
    11: '6.1"',
    12: '6.7"',
    13: '6.1"',
    14: '6.7"',
    15: '6.1"',
    16: '6.7"',
    17: '6.3"',
    18: '6.9"',
    19: '6.3"',
    20: '6.9"',
    21: '6.3"',
    22: '6.0"',
    23: '6.2"',
    24: '6.1"',
    25: '6.4"',
    26: '6.3"',
    27: '6.2"',
    28: '6.1"',
    29: '6.7"',
    30: '6.3"',
    31: '6.3"',
    32: '6.8"',
    33: '6.3"',
    34: '6.3"',
    35: '6.3"',
    36: '6.8"',
    37: '6.3"',
    38: 'Mixed',
    39: 'Mixed',
    40: 'Mixed',
    41: 'Mixed',
    42: 'Mixed',
    43: 'Mixed',
    44: 'Mixed',
    45: 'Mixed',
    46: 'Mixed',
    47: '6.7"',
    48: '6.7"',
    49: '6.67"',
    50: '6.6"',
    51: '6.6"',
    52: '6.7"',
    53: 'Mixed',
    54: '6.53"',
    55: 'Mixed',
    56: '6.3"',
    57: 'Mixed',
    58: '6.5"',
    59: '6.43"',
    60: 'Mixed',
    61: '6.4"',
    62: 'Mixed',
    63: '6.43"',
    64: '6.22"',
    65: 'Mixed',
    66: '5.5"',
    67: '5.5"',
    68: '5.0"',
    69: '5.6"',
    70: '5.0"',
    71: 'Mixed',
    72: '5.7"',
    73: '5.2"',
    74: '6.4"',
    75: '6.43"',
    76: '5.45"',
    77: '6.59"',
    78: '5.7"',
    79: '6.5"',
    80: '6.26"',
    81: '5.84"',
    82: '5.9"',
    83: '5.45"',
    84: '6.5"',
    85: 'Mixed',
    86: '6.39"',
    87: '6.7"',
    88: 'Mixed',
    89: '6.7"',
    90: 'Mixed',
    91: 'Mixed',
    92: '6.53"',
    93: '6.43"',
    94: '6.62"',
    95: '6.88"',
    96: 'Unknown',
    97: '6.75"',
    98: '6.44"',
    99: '6.5"',
    100: '6.6"',
    101: '6.5"',
    102: '6.6"',
    103: '6.6"',
    104: '6.0"',
    105: '6.7"',
    106: '6.7"',
    107: '6.59"',
    108: '6.7"',
    109: '5.2"',
    110: '6.78"',
    111: '6.43"',
    112: 'Mixed',
    113: '6.55"',
    114: '6.82"',
    115: '6.9"',
    116: '6.8"',
    117: 'Unknown',
    118: 'Mixed',
    119: '6.7"',
    120: '6.95"',
    121: '6.78"',
    122: 'Unknown',
    123: '6.67"',
    124: '6.5"',
    125: '6.88"',
    126: '6.72"',
    127: '6.88"',
    128: '6.78"',
    129: 'Unknown',
    130: '6.55"'
}

def clean_model_name(brand, model_text):
    model_text = model_text.strip()
    
    # If model starts with SAM A... / SAM S... / SAM M... add Galaxy if appropriate or standard Samsung brand
    if brand == 'Samsung':
        # E.g., A34 5G -> Samsung Galaxy A34 5G
        if re.match(r'^(A|S|M|F|Z|Note)\d+', model_text, re.I):
            return f"Samsung Galaxy {model_text}"
        elif not model_text.lower().startswith('samsung'):
            return f"Samsung {model_text}"
        return model_text
    elif brand == 'iPhone':
        if not model_text.lower().startswith('iphone'):
            return f"iPhone {model_text}"
        return model_text
    elif brand == 'Pixel':
        if not model_text.lower().startswith('pixel') and not model_text.lower().startswith('google pixel'):
            return f"Pixel {model_text}"
        return model_text
    elif brand:
        if not model_text.lower().startswith(brand.lower()):
            return f"{brand} {model_text}"
        return model_text
    
    return model_text

def parse_box_models(b_id, raw_text):
    if not raw_text or raw_text in ['None', '—']:
        return [], f"Box {b_id:03d}"
    
    lines = raw_text.split('\n')
    models = []
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        current_brand = ""
        
        # Check if line has a explicit brand prefix like "Redmi:", "Huawei:", "OPPO:", "Realme:", "Vivo:", "Tecno:", "Infinix:"
        brand_match = re.match(r'^([A-Za-z0-9\+\s]+):\s*(.*)$', line)
        if brand_match:
            b_prefix = brand_match.group(1).strip()
            rest = brand_match.group(2).strip()
            upper_b = b_prefix.upper()
            current_brand = BRAND_ABBREVIATIONS.get(upper_b, b_prefix)
            line = rest
        
        parts = [p.strip() for p in line.split('/') if p.strip()]
        
        line_brand = current_brand
        
        for part in parts:
            first_word = part.split()[0].upper() if part.split() else ''
            
            if first_word in BRAND_ABBREVIATIONS:
                line_brand = BRAND_ABBREVIATIONS[first_word]
                part_model = ' '.join(part.split()[1:])
            elif first_word in ['IPHONE', 'SAMSUNG', 'REDMI', 'OPPO', 'VIVO', 'ONEPLUS', 'POCO', 'XIAOMI', 'REALME', 'TECNO', 'HUAWEI', 'HONOR', 'INFINIX', 'MOTOROLA', 'NOKIA', 'ZTE', 'PIXEL']:
                line_brand = 'iPhone' if first_word == 'IPHONE' else ('OnePlus' if first_word == 'ONEPLUS' else ('POCO' if first_word == 'POCO' else ('Pixel' if first_word == 'PIXEL' else first_word.capitalize())))
                part_model = ' '.join(part.split()[1:])
            else:
                part_model = part
            
            full_model = clean_model_name(line_brand, part_model)
            full_model = re.sub(r'\s+', ' ', full_model).strip()
            
            if full_model and full_model not in models:
                models.append(full_model)
                
    title = models[0] if models else f"Box {b_id:03d}"
    return models, title

boxes_data = []

for r in range(2, 132):
    b_id = int(sheet.cell(r, 1).value)
    b_num = f"BOX {b_id:03d}"
    raw_val = sheet.cell(r, 2).value
    raw_text = str(raw_val).strip() if raw_val is not None else ""
    
    models, title = parse_box_models(b_id, raw_text)
    display_size = DISPLAY_SIZES.get(b_id, "Unknown")
    
    boxes_data.append({
        "id": f"box-{b_id:03d}",
        "boxNumber": b_num,
        "displaySize": display_size,
        "title": title,
        "rawText": raw_text,
        "compatibleModels": models
    })

total_relationships = sum(len(b["compatibleModels"]) for b in boxes_data)
all_models_set = set(m for b in boxes_data for m in b["compatibleModels"])

json_output = {
    "version": "2.0.0-130box",
    "lastUpdated": datetime.datetime.now(datetime.timezone.utc).isoformat(),
    "totalBoxes": len(boxes_data),
    "totalRelationships": total_relationships,
    "uniqueModels": len(all_models_set),
    "boxes": boxes_data
}

# Write src/data/screenguards.json
json_filepath = 'src/data/screenguards.json'
with open(json_filepath, 'w', encoding='utf-8') as f:
    json.dump(json_output, f, indent=2, ensure_ascii=False)

print(f"Generated {json_filepath}: {len(boxes_data)} boxes, {total_relationships} model relationships, {len(all_models_set)} unique models.")

# Generate seed-data.sql
sql_lines = [
    "-- UZEE TECH SCREENGUARD FINDER — 130-BOX AUTHORITATIVE PRODUCTION SEED SQL",
    f"-- Generated on {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    "-- DO NOT MERGE WITH OLD 106/263 DATASET. THIS TOTALLY REPLACES THE PRODUCTION DATASET.",
    "",
    "BEGIN;",
    "",
    "-- Clear existing production data safely",
    "TRUNCATE TABLE models CASCADE;",
    "TRUNCATE TABLE boxes CASCADE;",
    "",
    "-- Insert 130 physical boxes",
    "INSERT INTO boxes (id, box_number, title, display_size, raw_text) VALUES"
]

box_rows = []
for b in boxes_data:
    b_id = b["id"]
    b_num = b["boxNumber"]
    title_esc = b["title"].replace("'", "''")
    size_esc = b["displaySize"].replace("'", "''")
    raw_esc = b["rawText"].replace("'", "''")
    box_rows.append(f"  ('{b_id}', '{b_num}', '{title_esc}', '{size_esc}', '{raw_esc}')")

sql_lines.append(",\n".join(box_rows) + ";")
sql_lines.append("")
sql_lines.append("-- Insert compatible model relationships")
sql_lines.append("INSERT INTO models (box_id, model_name) VALUES")

model_rows = []
for b in boxes_data:
    b_id = b["id"]
    for m in b["compatibleModels"]:
        m_esc = m.replace("'", "''")
        model_rows.append(f"  ('{b_id}', '{m_esc}')")

if model_rows:
    sql_lines.append(",\n".join(model_rows) + ";")
else:
    sql_lines.append("-- No models inserted")

sql_lines.append("")
sql_lines.append("COMMIT;")
sql_lines.append("")

sql_filepath = 'seed-data.sql'
with open(sql_filepath, 'w', encoding='utf-8') as f:
    f.write("\n".join(sql_lines))

print(f"Generated {sql_filepath} with {len(boxes_data)} boxes and {len(model_rows)} model records.")
